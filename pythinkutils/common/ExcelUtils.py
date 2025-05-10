# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook

from pythinkutils.common.StringUtils import *
from pythinkutils.common.object2json import *

class ExcelUtils(object):
    @classmethod
    def excel_to_list(cls, szExcel):
        szExcel = szExcel.strip()
        lstLine = szExcel.split("\n")

        lstHeader = None
        lstRet = []
        nLine = 0
        for szLine in lstLine:
            if 0 == nLine:
                lstHeader = szLine.strip().split("\t")
                # g_logger.info(obj2json(lstHeader))
            else:
                dictItem = {}
                lstItem = szLine.strip().split("\t")
                nPos = 0
                for szItem in lstItem:
                    if nPos >= len(lstHeader):
                        continue

                    dictItem[lstHeader[nPos].strip()] = szItem.strip()
                    nPos += 1

                lstRet.append(dictItem)

            # g_logger.info(szLine)
            nLine += 1

        return lstRet

    @classmethod
    def read_excel(cls, szFile, szSheet, nRow, nCol, workbook = None) -> str:
        if workbook is None:
            workbook = load_workbook(szFile)

        sheet = workbook[szSheet]
        return str(sheet.cell(row=nRow, column=nCol).value)

    @classmethod
    def write_excel(cls, szFile, szSheet, nRow, nCol, szVal, workbook=None) -> str:
        if workbook is None:
            workbook = load_workbook(szFile)

        if is_empty_string(szSheet):
            sheet = workbook.active
        else:
            sheet = workbook[szSheet]

        sheet.cell(row=nRow, column=nCol, value=szVal)
        return str(sheet.cell(row=nRow, column=nCol).value)

    @classmethod
    def _dictitem_2_str(cls, item):
        if str == type(item):
            return item
        elif dict == type(item):
            return obj2json(item)
        elif list == type(item):
            return obj2json(item)
        else:
            return item

    @classmethod
    def dictlist_2_xls(cls, lstData, szFilePath, szSheet, lstHeader):
        wb = Workbook()
        if is_empty_string(szSheet):
            sheet = wb.active
        else:
            sheet = wb.create_sheet(szSheet)

        # make hearer
        for nCol in range(len(lstHeader)):
            cls.write_excel(None, szSheet, 1, nCol+1, lstHeader[nCol], wb)

        # make data
        for nRow in range(len(lstData)):
            for nCol in range(len(lstHeader)):
                if lstHeader[nCol] in lstData[nRow].keys():
                    szVal = cls._dictitem_2_str(lstData[nRow][lstHeader[nCol]])
                    cls.write_excel(None, szSheet, nRow+2, nCol+1, szVal, wb)
                else:
                    continue

        wb.save(szFilePath)

    @classmethod
    def all_sheet(cls, szFile) -> list:
        workbook = load_workbook(szFile)
        return workbook.sheetnames


# def main():
#     from pythinkutils.common.log import g_logger
#     from pythinkutils.common.object2json import obj2json
#
#     szExcel = '''
#         id	name	icon	des	lv	type	baseProp	spProp	s5Prop	s10Prop	s15Prop	suitBox
#         1	武器	icon_equip	破旧不堪的武器	1	1	1:106:10		1:100:0	6:1:0	12:1:0
#         2	头盔	icon_equip	破旧不堪的头盔	1	2	1:71:10		1:100:0	6:1:0	12:1:0
#         3	衣服	icon_equip	破旧不堪的衣服	1	3	2:805:10		2:1000:0	7:2:0	11:2:0
#         4	鞋子	icon_equip	破旧不堪的鞋子	1	4	2:536:10		2:1000:0	7:2:0	11:2:0
#         5	武器	icon_equip	破旧不堪的武器	2	1	1:237:10		1:100:0	6:1:0	12:1:0
#         6	头盔	icon_equip	破旧不堪的头盔	2	2	1:158:10		1:100:0	6:1:0	12:1:0
#         7	衣服	icon_equip	破旧不堪的衣服	2	3	2:1783:10		2:1000:0	7:2:0	11:2:0
#         8	鞋子	icon_equip	破旧不堪的鞋子	2	4	2:1189:10		2:1000:0	7:2:0	11:2:0
#         9	武器	icon_equip	破旧不堪的武器	3	1	1:703:10		1:100:0	6:1:0	12:1:0
#         10	头盔	icon_equip	破旧不堪的头盔	3	2	1:469:10		1:100:0	6:1:0	12:1:0
#         11	衣服	icon_equip	破旧不堪的衣服	3	3	2:5282:10		2:1000:0	7:2:0	11:2:0
#         12	鞋子	icon_equip	破旧不堪的鞋子	3	4	2:3522:10		2:1000:0	7:2:0	11:2:0
#         13	武器	icon_equip	破旧不堪的武器	4	1	1:1552:10|6:3:10		1:100:0	6:1:0	12:1:0
#         14	头盔	icon_equip	破旧不堪的头盔	4	2	1:1034:10|6:3:10		1:100:0	6:1:0	12:1:0
#         15	衣服	icon_equip	破旧不堪的衣服	4	3	2:11644:10|7:3:10		2:1000:0	7:2:0	11:2:0
#         16	鞋子	icon_equip	破旧不堪的鞋子	4	4	2:7762:10|7:3:10		2:1000:0	7:2:0	11:2:0
#         17	武器	icon_equip	破旧不堪的武器	5	1	1:2386:10|6:4:10		1:100:0	6:1:0	12:1:0
#         18	头盔	icon_equip	破旧不堪的头盔	5	2	1:1590:10|6:4:10		1:100:0	6:1:0	12:1:0
#         19	衣服	icon_equip	破旧不堪的衣服	5	3	2:17905:10|7:4:10		2:1000:0	7:2:0	11:2:0
#         20	鞋子	icon_equip	破旧不堪的鞋子	5	4	2:11937:10|7:4:10		2:1000:0	7:2:0	11:2:0
#         21	武器	icon_equip	破旧不堪的武器	6	1	1:2386:10|6:5:10		1:100:0	6:1:0	12:1:0
#         22	头盔	icon_equip	破旧不堪的头盔	6	2	1:1590:10|6:5:10		1:100:0	6:1:0	12:1:0
#         23	衣服	icon_equip	破旧不堪的衣服	6	3	2:17905:10|7:5:10		2:1000:0	7:2:0	11:2:0
#         24	鞋子	icon_equip	破旧不堪的鞋子	6	4	2:11937:10|7:5:10		2:1000:0	7:2:0	11:2:0
#         25	武器	icon_equip	破旧不堪的武器	6	1	1:2386:10|6:5:10		1:100:0	6:1:0	12:1:0	11
#         26	头盔	icon_equip	破旧不堪的头盔	6	2	1:1590:10|6:5:10		1:100:0	6:1:0	12:1:0	11
#         27	衣服	icon_equip	破旧不堪的衣服	6	3	2:17905:10|7:5:10		2:1000:0	7:2:0	11:2:0	11
#         28	鞋子	icon_equip	破旧不堪的鞋子	6	4	2:11937:10|7:5:10		2:1000:0	7:2:0	11:2:0	11
#     '''
#
#     lstData = ExcelUtils.excel_to_list(szExcel)
#     g_logger.info(obj2json(lstData).encode('utf-8').decode('unicode_escape'))
#
# if __name__ == '__main__':
#     main()